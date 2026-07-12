"""Excel / CSV 解析与生成工具

Excel 基于 openpyxl，无需 pandas；CSV 基于标准库 csv。
列结构: time / value / label（label 可选）
"""
import csv
import io
import logging
from datetime import datetime
from typing import Any, Optional

logger = logging.getLogger(__name__)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.utils.response import APIError


# 表头别名映射（大小写不敏感、支持中英文）
TIME_HEADERS = {"time", "时间", "timestamp", "date", "日期"}
VALUE_HEADERS = {"value", "值", "数据", "数值", "观测值"}
LABEL_HEADERS = {"label", "标签", "标注", "备注"}


def parse_excel(file_bytes: bytes) -> dict:
    """解析 Excel 文件

    Returns:
        {
            "series_data": [{"time": str, "value": float, "label": str}, ...],
            "point_count": int,
            "warnings": [str, ...],
        }
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise APIError(f"Excel 文件解析失败: {e}", status_code=400)

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise APIError("Excel 文件至少需要表头 + 1 行数据", status_code=400)

    # 识别表头列索引
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    time_idx = _find_header(header, TIME_HEADERS)
    value_idx = _find_header(header, VALUE_HEADERS)
    label_idx = _find_header(header, LABEL_HEADERS)

    if value_idx is None:
        raise APIError("Excel 表头未找到 value 列（支持: value/值/数据/数值/观测值）", status_code=400)
    if time_idx is None:
        raise APIError("Excel 表头未找到 time 列（支持: time/时间/日期/date/timestamp）", status_code=400)

    series_data = []
    warnings = []
    for i, row in enumerate(rows[1:], start=2):  # Excel 行号从 2 开始（表头是第1行）
        if row is None:
            continue
        # 跳过全空行
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue

        time_val = row[time_idx] if time_idx < len(row) else None
        raw_val = row[value_idx] if value_idx < len(row) else None
        label_val = row[label_idx] if (label_idx is not None and label_idx < len(row)) else None

        # value 必须为数值
        num_val = _to_float(raw_val)
        if num_val is None:
            warnings.append(f"第 {i} 行 value 非数值（{raw_val}），已跳过")
            continue

        # time 缺失时按序号生成
        time_str = _to_time_str(time_val, fallback_idx=len(series_data))
        label_str = str(label_val).strip() if label_val is not None else ""

        series_data.append({"time": time_str, "value": num_val, "label": label_str})

    if len(series_data) < 3:
        raise APIError(f"有效数据点仅 {len(series_data)} 个，至少需要 3 个（模型要求）", status_code=400)

    return {
        "series_data": series_data,
        "point_count": len(series_data),
        "warnings": warnings,
    }


def generate_template(file_bytes: bytes = None) -> bytes:
    """生成 Excel 模板文件（含示例数据）

    返回 xlsx 文件的字节内容。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "数据集"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    # 写表头
    headers = ["time", "value", "label"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 示例数据（12 个月，含上升趋势 + 季节波动）
    sample_data = [
        ("2024-01", 120.5, ""),
        ("2024-02", 135.2, ""),
        ("2024-03", 142.8, "旺季"),
        ("2024-04", 128.4, ""),
        ("2024-05", 138.9, ""),
        ("2024-06", 145.6, ""),
        ("2024-07", 152.3, ""),
        ("2024-08", 148.7, ""),
        ("2024-09", 155.1, ""),
        ("2024-10", 162.4, "促销月"),
        ("2024-11", 158.9, ""),
        ("2024-12", 168.2, ""),
    ]
    for row_idx, (t, v, l) in enumerate(sample_data, start=2):
        ws.cell(row=row_idx, column=1, value=t)
        ws.cell(row=row_idx, column=2, value=v)
        ws.cell(row=row_idx, column=3, value=l)

    # 列宽
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16

    # 第二个 sheet：填写说明
    ws2 = wb.create_sheet("填写说明")
    instructions = [
        ["列名", "必填", "说明"],
        ["time", "是", "时间标签。支持格式: 2024-01 / 2024-01-15 / 2024 / 2024/1 / t1。缺失时按行序自动生成 t1, t2..."],
        ["value", "是", "观测值，必须为数值。非数值行将被跳过并记录警告。"],
        ["label", "否", "数据点标注（如'促销月'、'异常'、'设备检修'），用于图表标注。可留空。"],
        ["", "", ""],
        ["注意事项", "", ""],
        ["1. 第一行为表头，固定 time / value / label（大小写不敏感，支持中文 时间/值/标签）", ""],
        ["2. 至少需要 3 个有效数据点", ""],
        ["3. 建议数据点数 >= 10 以获得更好预测效果", ""],
        ["4. label 列可整列留空", ""],
    ]
    for row_idx, row in enumerate(instructions, start=1):
        for col_idx, val in enumerate(row, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1 or (len(row) > 0 and row[0] == "列名"):
                cell.font = header_font
                cell.fill = header_fill
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 80

    # 导出为字节
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_dataset_excel(name: str, frequency: str, unit: str, series_data: list,
                         forecasts: list = None, future_times: list = None,
                         quantiles: dict = None, actuals: list = None,
                         metrics: dict = None) -> bytes:
    """导出数据集（含可选预测结果）为 Excel

    Args:
        series_data: 原始数据 [{time, value, label}]
        forecasts: 预测点预测值列表（可选）
        future_times: 预测时间标签列表（可选）
        quantiles: 分位数预测（可选）
        actuals: 回测实际值列表（可选，仅回测模式）
        metrics: 回测误差指标（可选）

    Returns:
        xlsx 字节内容
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"

    # 表头
    headers = ["time", "value", "label"]
    has_forecast = forecasts is not None and len(forecasts) > 0
    has_actuals = actuals is not None and len(actuals) > 0
    if has_forecast:
        headers += ["type", "forecast", "p10", "p90"]
        if has_actuals:
            headers += ["actual", "error"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill

    # 原始数据
    for row_idx, point in enumerate(series_data, start=2):
        ws.cell(row=row_idx, column=1, value=point.get("time", ""))
        ws.cell(row=row_idx, column=2, value=point.get("value", 0))
        ws.cell(row=row_idx, column=3, value=point.get("label", ""))
        if has_forecast:
            ws.cell(row=row_idx, column=4, value="history")

    # 预测数据
    if has_forecast:
        p10 = (quantiles or {}).get("0.1", [])
        p90 = (quantiles or {}).get("0.9", [])
        start_row = len(series_data) + 2
        for i, fc in enumerate(forecasts):
            r = start_row + i
            ws.cell(row=r, column=1, value=future_times[i] if i < len(future_times) else f"t{i+1}")
            ws.cell(row=r, column=2, value="")
            ws.cell(row=r, column=3, value="")
            ws.cell(row=r, column=4, value="forecast")
            ws.cell(row=r, column=5, value=fc)
            ws.cell(row=r, column=6, value=p10[i] if i < len(p10) else "")
            ws.cell(row=r, column=7, value=p90[i] if i < len(p90) else "")
            if has_actuals and i < len(actuals):
                ws.cell(row=r, column=8, value=actuals[i] if actuals[i] is not None else "")
                # F-4: 防御 actuals[i] 为 None
                ws.cell(row=r, column=9, value=round(actuals[i] - fc, 4) if actuals[i] is not None else "")

    # 元信息 sheet
    ws2 = wb.create_sheet("元信息")
    meta = [
        ["名称", name],
        ["频率", frequency],
        ["单位", unit],
        ["数据点数", len(series_data)],
    ]
    if has_forecast:
        meta.append(["预测点数", len(forecasts)])
    if has_actuals and metrics:
        meta.append(["--- 回测指标 ---", ""])
        for mk, mv in metrics.items():
            meta.append([f"  {mk.upper()}", mv])
    for row_idx, (k, v) in enumerate(meta, start=1):
        ws2.cell(row=row_idx, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=row_idx, column=2, value=v)
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _find_header(header: list, candidates: set) -> Any:
    """在表头中查找匹配的列索引"""
    for idx, h in enumerate(header):
        if h and h in candidates:
            return idx
    return None


def _to_float(val) -> Optional[float]:
    """尝试转为 float"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).strip()
        if not s:
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


def _to_time_str(val, fallback_idx: int) -> str:
    """将 time 单元格值转为字符串"""
    if val is None or (isinstance(val, str) and not val.strip()):
        return f"t{fallback_idx + 1}"
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    # 日期对象
    if hasattr(val, "strftime"):
        try:
            return val.strftime("%Y-%m-%d")
        except Exception:
            logger.debug(f"[excel_parser] strftime 失败 val={val!r}")
    return str(val).strip()


# ==================== CSV 支持 ====================

def parse_csv(file_bytes: bytes, encoding: str = "utf-8-sig") -> dict:
    """解析 CSV 文件

    自动处理 BOM；列结构与 Excel 一致: time / value / label
    返回结构与 parse_excel 相同。
    """
    # 尝试多种编码
    text = None
    for enc in [encoding, "utf-8-sig", "utf-8", "gbk", "gb18030"]:
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise APIError("CSV 文件编码解析失败，请保存为 UTF-8 编码", status_code=400)

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if len(rows) < 2:
        raise APIError("CSV 文件至少需要表头 + 1 行数据", status_code=400)

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    time_idx = _find_header(header, TIME_HEADERS)
    value_idx = _find_header(header, VALUE_HEADERS)
    label_idx = _find_header(header, LABEL_HEADERS)

    if value_idx is None:
        raise APIError("CSV 表头未找到 value 列（支持: value/值/数据/数值/观测值）", status_code=400)
    if time_idx is None:
        raise APIError("CSV 表头未找到 time 列（支持: time/时间/日期/date/timestamp）", status_code=400)

    series_data = []
    warnings = []
    for i, row in enumerate(rows[1:], start=2):
        if not row or all(not (c or "").strip() for c in row):
            continue

        time_val = row[time_idx] if time_idx < len(row) else None
        raw_val = row[value_idx] if value_idx < len(row) else None
        label_val = row[label_idx] if (label_idx is not None and label_idx < len(row)) else None

        num_val = _to_float(raw_val)
        if num_val is None:
            warnings.append(f"第 {i} 行 value 非数值（{raw_val}），已跳过")
            continue

        time_str = _to_time_str(time_val, fallback_idx=len(series_data))
        label_str = str(label_val).strip() if label_val is not None else ""

        series_data.append({"time": time_str, "value": num_val, "label": label_str})

    if len(series_data) < 3:
        raise APIError(f"有效数据点仅 {len(series_data)} 个，至少需要 3 个（模型要求）", status_code=400)

    return {
        "series_data": series_data,
        "point_count": len(series_data),
        "warnings": warnings,
    }


def generate_csv_template() -> bytes:
    """生成 CSV 模板文件（含示例数据）"""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["time", "value", "label"])
    sample = [
        ("2024-01", "120.5", ""),
        ("2024-02", "135.2", ""),
        ("2024-03", "142.8", "旺季"),
        ("2024-04", "128.4", ""),
        ("2024-05", "138.9", ""),
        ("2024-06", "145.6", ""),
        ("2024-07", "152.3", ""),
        ("2024-08", "148.7", ""),
        ("2024-09", "155.1", ""),
        ("2024-10", "162.4", "促销月"),
        ("2024-11", "158.9", ""),
        ("2024-12", "168.2", ""),
    ]
    for row in sample:
        writer.writerow(row)
    # 加 BOM 便于 Excel 正确识别 UTF-8
    return b"\xef\xbb\xbf" + output.getvalue().encode("utf-8")


def export_dataset_csv(name: str, frequency: str, unit: str, series_data: list,
                       forecasts: list = None, future_times: list = None,
                       quantiles: dict = None, actuals: list = None,
                       metrics: dict = None) -> bytes:
    """导出数据集（含可选预测结果）为 CSV"""
    output = io.StringIO()
    writer = csv.writer(output)

    has_forecast = forecasts is not None and len(forecasts) > 0
    has_actuals = actuals is not None and len(actuals) > 0
    header = ["time", "value", "label"]
    if has_forecast:
        header += ["type", "forecast", "p10", "p90"]
        if has_actuals:
            header += ["actual", "error"]
    writer.writerow(header)

    for point in series_data:
        row = [point.get("time", ""), point.get("value", 0), point.get("label", "")]
        if has_forecast:
            row.append("history")
            row += ["", "", ""]
            if has_actuals:
                row += ["", ""]
        writer.writerow(row)

    if has_forecast:
        p10 = (quantiles or {}).get("0.1", [])
        p90 = (quantiles or {}).get("0.9", [])
        for i, fc in enumerate(forecasts):
            ft = future_times[i] if i < len(future_times) else f"t{i+1}"
            row = [
                ft, "", "", "forecast", fc,
                p10[i] if i < len(p10) else "",
                p90[i] if i < len(p90) else "",
            ]
            if has_actuals and i < len(actuals):
                # F-4: 防御 actuals[i] 为 None
                row.append(actuals[i] if actuals[i] is not None else "")
                row.append(round(actuals[i] - fc, 4) if actuals[i] is not None else "")
            elif has_actuals:
                row += ["", ""]
            writer.writerow(row)

    # 回测指标追加在末尾
    if has_actuals and metrics:
        writer.writerow([])
        writer.writerow(["# 回测误差指标"])
        for mk, mv in metrics.items():
            writer.writerow([f"#{mk.upper()}", mv])

    return b"\xef\xbb\xbf" + output.getvalue().encode("utf-8")


def parse_file(file_bytes: bytes, filename: str) -> dict:
    """根据文件扩展名自动选择解析器

    支持: .xlsx / .xls / .csv
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "csv":
        return parse_csv(file_bytes)
    else:
        # xlsx / xls 都用 openpyxl
        return parse_excel(file_bytes)

